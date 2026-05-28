"""Excel importer for test steps."""

from pathlib import Path
from openpyxl import load_workbook
from typing import Optional

from core.models import ActionType
from core.security import sanitize_step
from core.exceptions import ReportError

def list_excel_scenarios(xlsx_path: Path) -> list[dict]:
    """List scenarios from an Excel file."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        # Assume first sheet
        ws = wb.active
        
        scenarios = []
        current_scenario = None
        
        for row in ws.iter_rows(min_row=2, values_only=True): # Skip header
            # Row mapping: 
            # 0: Scenario ID (A)
            # 1: TC ID (B)
            # 2: TC Name (C)
            # 3: TC Description (D)
            # ...
            # 6: Step Description (G)
            
            if len(row) < 7: continue
            
            scen_id = row[0]
            tc_name = row[2]
            
            if scen_id: # New scenario
                current_scenario = {
                    "id": str(scen_id),
                    "name": str(tc_name) if tc_name else f"Scenario {scen_id}",
                    "description": str(row[3]) if row[3] else "",
                    "step_count": 0
                }
                scenarios.append(current_scenario)
                
            if current_scenario and row[6]: # Has step description
                current_scenario["step_count"] += 1
                
        return scenarios
        
    except Exception as e:
        raise ReportError(f"Failed to read Excel scenarios: {e}")

def import_excel_steps(xlsx_path: Path, scenario_id: Optional[str] = None) -> list[dict]:
    """Parse steps from Excel."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        
        steps = []
        in_target_scenario = False if scenario_id else True
        sequence = 1
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 9: continue
            
            curr_scen_id = row[0]
            if curr_scen_id:
                if scenario_id:
                    in_target_scenario = (str(curr_scen_id) == scenario_id)
                else:
                    in_target_scenario = True
            
            if not in_target_scenario:
                continue
                
            step_desc = row[6]
            expected = row[7]
            
            if not step_desc: continue
            
            step_str = str(step_desc)
            action = "navigate" if "pre-condition" in str(row[5] or "").lower() else "assert_text"
            
            # Create a raw dict
            raw_step = {
                "action": action,
                "selector": "",
                "value": step_str,
                "description": step_str,
                "expected": expected,
                "sequence": sequence,
                "is_sensitive": False
            }
            
            # Sanitize it
            sanitized = sanitize_step(raw_step)
            steps.append(sanitized)
            sequence += 1
            
        return steps
        
    except Exception as e:
        raise ReportError(f"Failed to import Excel steps: {e}")
