"""Excel report generation."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.models import Run, Result
from core.exceptions import ReportError

def generate_excel_report(run: Run, results: list[Result], output_path: Path) -> Path:
    """Generate an Excel results report."""
    try:
        wb = Workbook()
        
        # --- Sheet 1: Summary ---
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Title
        ws_summary.merge_cells('A1:F1')
        title_cell = ws_summary['A1']
        title_cell.value = "QA Test Run Report"
        title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="0070F3", end_color="0070F3", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Info Table Data
        info_data = [
            ("Test Name", run.test_name),
            ("Run ID", run.id),
            ("Status", run.status.value.upper()),
            ("Started", run.started_at or run.created_at),
            ("Duration", f"{run.duration_seconds:.1f}s" if run.duration_seconds else "-"),
            ("Consultant", run.consultant or "-"),
            ("Pod", run.pod or "-"),
            ("Pass Rate", f"{(run.passed_count / run.step_count * 100) if run.step_count else 0:.1f}%"),
            ("Total Steps", str(run.step_count)),
            ("Passed", str(run.passed_count)),
            ("Failed", str(run.failed_count)),
            ("Skipped", str(run.step_count - run.passed_count - run.failed_count))
        ]
        
        row_num = 3
        for label, value in info_data:
            ws_summary.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            val_cell = ws_summary.cell(row=row_num, column=2, value=value)
            
            if label == "Status":
                if value == "PASSED":
                    val_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    val_cell.font = Font(color="006100", bold=True)
                elif value == "FAILED":
                    val_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    val_cell.font = Font(color="9C0006", bold=True)
            row_num += 1
            
        # --- Sheet 2: Steps ---
        ws_steps = wb.create_sheet(title="Steps")
        
        headers = ["Step #", "Action", "Description", "Selector", "Expected", "Actual", "Status", "Duration (ms)", "Screenshot"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_steps.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0070F3", end_color="0070F3", fill_type="solid")
        
        ws_steps.freeze_panes = "A2"
        ws_steps.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        
        for i, res in enumerate(results, 2):
            ws_steps.cell(row=i, column=1, value=res.sequence)
            ws_steps.cell(row=i, column=2, value=res.action.value)
            ws_steps.cell(row=i, column=3, value=res.description)
            ws_steps.cell(row=i, column=4, value=res.selector)
            ws_steps.cell(row=i, column=5, value=res.expected_value)
            
            # Value redaction not directly in result, but we can check if it says REDACTED in DB or assume step is_sensitive handled
            val = res.actual_value
            if val and "REDACTED" in val:
                val = "[REDACTED]"
            ws_steps.cell(row=i, column=6, value=val)
            
            status_cell = ws_steps.cell(row=i, column=7, value=res.status.upper())
            if res.status == 'passed':
                status_cell.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
            elif res.status == 'failed':
                status_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            elif res.status == 'skipped':
                status_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
            ws_steps.cell(row=i, column=8, value=res.duration_ms)
            
            ss = res.screenshot_path.split('\\')[-1].split('/')[-1] if res.screenshot_path else ""
            ws_steps.cell(row=i, column=9, value=ss)
            
        # --- Sheet 3: Metadata ---
        ws_meta = wb.create_sheet(title="Metadata")
        meta_data = [
            ("run_id", run.id),
            ("test_id", run.test_id),
            ("run_dir", run.run_dir),
            ("video_path", run.video_path),
            ("trace_path", run.trace_path),
            ("created_at", run.created_at),
            ("completed_at", run.completed_at),
            ("error_message", run.error_message)
        ]
        
        for r, (k, v) in enumerate(meta_data, 1):
            ws_meta.cell(row=r, column=1, value=k).font = Font(bold=True)
            ws_meta.cell(row=r, column=2, value=str(v) if v else "")

        # Auto-fit columns
        for sheet in wb.worksheets:
            for col_idx, column in enumerate(sheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 60)
                sheet.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        return output_path
        
    except Exception as e:
        raise ReportError(f"Failed to generate Excel report: {e}")
